"""엑셀 파일에서 배송 목록 읽기 (API 승인 전 / 비상용)
   - 쿠팡 WING 배송관리 > 엑셀 다운로드 (DeliveryList_*.xlsx)
   - 스마트스토어 발주/발송관리 엑셀 (비밀번호 걸린 파일도 지원)
"""
from __future__ import annotations

import io
import logging
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from .common import Shipment, normalize_phone

log = logging.getLogger("alimtalk.excel")


def _open(path: str | Path, password: str | None = None):
    p = Path(path)
    data = p.read_bytes()
    try:
        return load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        pass
    # 암호화(비밀번호) 파일
    import msoffcrypto
    of = msoffcrypto.OfficeFile(io.BytesIO(data))
    if not password:
        raise SystemExit(f"{p.name} 은 비밀번호가 걸린 파일입니다. --password 옵션으로 비밀번호를 넣어주세요.")
    of.load_key(password=password)
    buf = io.BytesIO()
    of.decrypt(buf)
    buf.seek(0)
    return load_workbook(buf, read_only=True, data_only=True)


def _rows_with_header(ws, must_have: str):
    """헤더 행(must_have 컬럼이 있는 행)을 찾아 dict 목록으로 반환"""
    rows = list(ws.iter_rows(values_only=True))
    hidx = None
    for i, r in enumerate(rows[:10]):
        if r and any(must_have in str(c) for c in r if c is not None):
            hidx = i
            break
    if hidx is None:
        raise SystemExit(f"'{must_have}' 컬럼이 있는 헤더 행을 찾지 못했습니다 (시트: {ws.title})")
    header = [str(c).strip() if c is not None else f"col{j}" for j, c in enumerate(rows[hidx])]
    out = []
    for r in rows[hidx + 1:]:
        if not r or all(c in (None, "") for c in r):
            continue
        out.append({header[j]: (r[j] if j < len(r) else None) for j in range(len(header))})
    return header, out


def _pick(row: dict, *cands: str, default=""):
    """후보 컬럼명 중 존재하는 첫 값을 반환 (부분일치 허용)"""
    for c in cands:
        if c in row and row[c] not in (None, ""):
            return row[c]
    for c in cands:
        for k, v in row.items():
            if c in k and v not in (None, ""):
                return v
    return default


# ------------------------------------------------------------ 쿠팡 DeliveryList
def coupang_excel(path, shop: str) -> list[Shipment]:
    wb = _open(path)
    ws = wb.worksheets[0]
    _, rows = _rows_with_header(ws, "주문번호")
    grouped = defaultdict(list)
    for r in rows:
        key = str(_pick(r, "묶음배송번호", "주문번호"))
        grouped[key].append(r)
    out = []
    for key, items in grouped.items():
        f = items[0]
        names = [str(_pick(it, "등록옵션명", "노출상품명(옵션명)")) for it in items]
        qty = sum(int(float(_pick(it, "구매수(수량)", "구매수", default=0) or 0)) for it in items)
        out.append(Shipment(
            shop=shop, channel="coupang_excel",
            order_id=str(_pick(f, "주문번호")),
            orderer_name=str(_pick(f, "구매자")),
            phone=normalize_phone(_pick(f, "구매자전화번호")),
            product_names=names, quantity=qty,
            courier=str(_pick(f, "택배사")),
            tracking_no=str(_pick(f, "운송장번호")),
            shipped_at=str(_pick(f, "출고일(발송일)", "주문시 출고예정일")),
            raw={"rows": items}))
    log.info("쿠팡 엑셀 %s → %d건", Path(path).name, len(out))
    return out


# ------------------------------------------------------------ 네이버 발주발송관리 엑셀
def naver_excel(path, shop: str, password: str | None = None) -> list[Shipment]:
    wb = _open(path, password)
    ws = wb.worksheets[0]
    _, rows = _rows_with_header(ws, "주문번호")
    grouped = defaultdict(list)
    for r in rows:
        grouped[str(_pick(r, "주문번호"))].append(r)
    out = []
    for oid, items in grouped.items():
        f = items[0]
        names = []
        for it in items:
            opt = str(_pick(it, "옵션정보", "옵션명", default="")).strip()
            names.append(opt or str(_pick(it, "상품명")))
        qty = sum(int(float(_pick(it, "수량", default=0) or 0)) for it in items)
        out.append(Shipment(
            shop=shop, channel="naver_excel", order_id=oid,
            orderer_name=str(_pick(f, "주문자명", "구매자명")),
            phone=normalize_phone(_pick(f, "주문자연락처1", "주문자연락처", "구매자연락처")),
            product_names=names, quantity=qty,
            courier=str(_pick(f, "택배사")),
            tracking_no=str(_pick(f, "송장번호", "운송장번호")),
            shipped_at=str(_pick(f, "발송일", "발송처리일")),
            raw={"rows": items}))
    log.info("네이버 엑셀 %s → %d건", Path(path).name, len(out))
    return out
