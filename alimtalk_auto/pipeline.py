"""수집 → 필터 → 발송(또는 엑셀 출력) → 이력 기록"""
from __future__ import annotations

import logging
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .aligo import Aligo, AligoError, send_sms
from .cleaner import product_label
from .common import Shipment, why_skip_today
from .state import State
from .store import SentLog

log = logging.getLogger("alimtalk.pipeline")
KST = timezone(timedelta(hours=9))


# ------------------------------------------------------------------ 수집
def collect(cfg: dict, shop_key: str, *, naver_excel=None, coupang_excel=None,
            password=None) -> list[Shipment]:
    shop = cfg["shops"][shop_key]
    ch = shop.get("channels", {})
    out: list[Shipment] = []

    if naver_excel or coupang_excel:            # 엑셀 모드
        from .excel_sources import naver_excel as ne, coupang_excel as ce
        if naver_excel:
            out += ne(naver_excel, shop_key, password)
        if coupang_excel:
            out += ce(coupang_excel, shop_key)
        return out

    since = datetime.now(KST).replace(hour=0, minute=0, second=0, microsecond=0)
    if not cfg.get("safety", {}).get("only_today", True):
        since -= timedelta(days=int(ch.get("naver", {}).get("lookback_days", 3)))

    if ch.get("naver", {}).get("enabled"):
        from .naver import NaverCommerce
        n = ch["naver"]
        try:
            out += NaverCommerce(n["client_id"], n["client_secret"]).fetch_shipments(
                shop_key, since, n.get("trigger", "DISPATCHED"), n.get("courier_map"))
        except Exception as e:
            log.error("네이버 수집 실패: %s", e)
    if ch.get("coupang", {}).get("enabled"):
        from .coupang import Coupang
        c = ch["coupang"]
        try:
            out += Coupang(c["vendor_id"], c["access_key"], c["secret_key"],
                           c.get("api_version", "v4")).fetch_shipments(
                shop_key, c.get("status", "DEPARTURE"), int(c.get("lookback_days", 7)))
        except Exception as e:
            log.error("쿠팡 수집 실패: %s", e)
    return out


# ------------------------------------------------------------------ 필터
def filter_shipments(cfg: dict, shipments: list[Shipment], sent: SentLog,
                     ignore_sent=False) -> tuple[list[Shipment], list[tuple[Shipment, str]]]:
    ok, skipped = [], []
    only_today = cfg.get("safety", {}).get("only_today", True)
    today = date.today().isoformat()
    for s in shipments:
        if not s.tracking_no:
            skipped.append((s, "운송장번호 없음")); continue
        if not s.phone:
            skipped.append((s, "주문자 연락처 없음")); continue
        if only_today and s.shipped_at and not s.shipped_at.startswith(today):
            skipped.append((s, f"오늘 발송건 아님({s.shipped_at[:10]})")); continue
        if not ignore_sent and sent.has(s.key()):
            skipped.append((s, "이미 발송됨")); continue
        ok.append(s)
    return ok, skipped


def build_values(s: Shipment, rules: dict | None) -> dict[str, str]:
    """알리고 템플릿 변수 → 값"""
    return {
        "고객명": s.orderer_name,
        "주문자": s.orderer_name,
        "주문상품명": product_label(s.product_names, rules),
        "상품명": product_label(s.product_names, rules),
        "택배사": s.courier,
        "운송장번호": s.tracking_no,
        "송장번호": s.tracking_no,
        "구매수량": str(s.quantity),
        "수량": str(s.quantity),
        "주문번호": s.order_id,
    }


# ------------------------------------------------------------------ 알리고 계정 / 포인트 알림
def aligo_account(cfg: dict, shop: dict) -> dict:
    """공통 aligo 설정 위에 상점별 설정을 덮어씀 (알리고 아이디가 상점마다 다른 경우 지원)"""
    acc = dict(cfg.get("aligo") or {})
    acc.update(shop.get("aligo") or {})
    for k in ("userid", "apikey", "sender"):
        if not acc.get(k):
            raise AligoError(f"알리고 {k} 설정이 없습니다 (config.yaml 확인)")
    return acc


def check_points(cfg: dict, shop_key: str, shop: dict, aligo: Aligo, need: int = 0) -> int | None:
    """잔여 알림톡 건수를 확인하고, 기준 이하이면 관리자에게 문자로 충전 안내를 보냄"""
    alert = cfg.get("point_alert") or {}
    left = aligo.remaining_alimtalk()
    if left is None:
        return None
    log.info("[%s] 알리고 잔여 알림톡 %s건", shop.get("display_name", shop_key), left)
    if not alert.get("enabled", True):
        return left
    threshold = int(alert.get("threshold", 300))
    if left > threshold and (need == 0 or left > need):
        return left

    st = State(cfg["_base_dir"])
    d = st.load()
    last = (d.get("notes") or {}).get(f"_point_alert_{shop_key}")
    today = date.today().isoformat()
    if last == today:
        log.info("포인트 부족 알림은 오늘 이미 보냈습니다")
        return left

    acc = aligo_account(cfg, shop)
    name = shop.get("display_name", shop_key)
    bank = (alert.get("bank_accounts") or {}).get(shop_key) or alert.get("bank_account") or ""
    msg = (f"[{name}] 알리고 잔액 부족\n"
           f"남은 알림톡 {left}건 (기준 {threshold}건)\n"
           f"오늘 발송 대상 {need}건\n")
    if bank:
        msg += f"충전 계좌: {bank}\n"
    msg += "충전: https://smartsms.aligo.in"
    try:
        send_sms(acc["userid"], acc["apikey"], acc["sender"],
                 alert.get("admin_phones") or [], msg, title=f"{name} 알리고 충전 필요")
        log.warning("포인트 부족 → 관리자에게 알림 문자 발송 (%s)", alert.get("admin_phones"))
        d.setdefault("notes", {})[f"_point_alert_{shop_key}"] = today
        st.save(d)
    except Exception as e:
        log.error("포인트 알림 문자 실패: %s", e)
    return left


# ------------------------------------------------------------------ 엑셀 출력 (알리고 업로드용 + 결과 리포트)
def write_aligo_xlsx(path: Path, aligo_vars: list[str], rows: list[dict]):
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    header = ["메시지 수신 휴대폰 번호"] + [f"#{{{v}}}" for v in aligo_vars]
    ws.append(header)
    for r in rows:
        ws.append([r["receiver"]] + [r["values"].get(v, "") for v in aligo_vars])
    for c in ws[1]:
        c.font = Font(bold=True)
    wb.save(path)


def write_report(path: Path, sent_rows: list[dict], skipped: list[tuple[Shipment, str]]):
    wb = Workbook(); ws = wb.active; ws.title = "발송"
    ws.append(["채널", "주문번호", "주문자", "수신번호", "상품", "택배사", "운송장", "발송일"])
    for r in sent_rows:
        s = r["shipment"]
        ws.append([s.channel, s.order_id, s.orderer_name, s.phone,
                   r["values"]["주문상품명"], s.courier, s.tracking_no, s.shipped_at])
    ws2 = wb.create_sheet("건너뜀")
    ws2.append(["채널", "주문번호", "주문자", "사유", "운송장", "발송일"])
    for s, why in skipped:
        ws2.append([s.channel, s.order_id, s.orderer_name, why, s.tracking_no, s.shipped_at])
    for w in (ws, ws2):
        for c in w[1]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="DDEBF7")
    wb.save(path)


# ------------------------------------------------------------------ 메인
def run_shop(cfg: dict, shop_key: str, *, dry_run=False, force=False,
             excel_only=False, naver_excel=None, coupang_excel=None,
             password=None, ignore_sent=False) -> dict:
    shop = cfg["shops"][shop_key]
    name = shop.get("display_name", shop_key)
    base = Path(cfg["_base_dir"])
    out_dir = base / "out"; out_dir.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M")

    # 1) 쉬는 날?
    why = why_skip_today(cfg)
    if why and not force:
        log.info("[%s] 오늘은 발송 안 함: %s  (--force 로 강제 실행 가능)", name, why)
        return {"skipped_day": why}

    # 2) 수집
    shipments = collect(cfg, shop_key, naver_excel=naver_excel,
                        coupang_excel=coupang_excel, password=password)
    log.info("[%s] 수집 %d건", name, len(shipments))

    # 3) 필터
    sent = SentLog(base / "sent.db")
    ok, skipped = filter_shipments(cfg, shipments, sent, ignore_sent=ignore_sent)
    for s, w in skipped:
        log.debug("건너뜀 %s %s: %s", s.channel, s.order_id, w)
    log.info("[%s] 발송 대상 %d건 / 건너뜀 %d건", name, len(ok), len(skipped))

    max_send = int(cfg.get("safety", {}).get("max_send_per_run", 300))
    if len(ok) > max_send:
        log.error("발송 대상(%d)이 max_send_per_run(%d)을 초과 → 중단. 설정을 확인하세요.", len(ok), max_send)
        return {"error": "too many"}

    # 4) 알리고 준비
    acc = aligo_account(cfg, shop)
    aligo = Aligo(acc["userid"], acc["apikey"], acc["sender"],
                  acc["senderkey"], acc["tpl_code"],
                  acc.get("failover", "N"), acc.get("testmode", "N"),
                  acc.get("template_text"))
    rules = shop.get("option_cleanup")
    items = []
    for s in ok:
        items.append({"receiver": s.phone, "recvname": s.orderer_name,
                      "values": build_values(s, rules), "shipment": s})

    # 4-1) 잔여 포인트 확인 (부족하면 관리자에게 충전 안내 문자)
    if not dry_run and not excel_only:
        try:
            check_points(cfg, shop_key, shop, aligo, need=len(items))
        except Exception as e:
            log.error("포인트 확인 중 오류(무시하고 진행): %s", e)

    # 5) 엑셀만 뽑기 (알리고 사이트에서 수동 업로드용)
    tpl_vars = None
    try:
        tpl_vars = aligo.variables if items else None
    except AligoError as e:
        log.warning("템플릿 조회 실패(%s) → 기본 변수로 엑셀 생성", e)
    if excel_only or dry_run:
        vars_ = tpl_vars or ["고객명", "주문상품명", "택배사", "운송장번호", "구매수량"]
        xp = out_dir / f"{shop_key}_알리고업로드_{stamp}.xlsx"
        write_aligo_xlsx(xp, vars_, items)
        log.info("알리고 업로드용 엑셀 저장: %s", xp)

    # 6) 발송
    result = {"sent": 0, "skipped": len(skipped)}
    if items and not excel_only:
        # 템플릿 변수 검증 (값이 없는 변수가 있으면 미리 에러)
        for it in items:
            aligo.render(it["values"])
        for i in range(0, len(items), 500):
            batch = items[i:i + 500]
            res = aligo.send_batch(batch, dry_run=dry_run)
            mid = (res.get("info") or {}).get("mid", "")
            for it in batch:
                sent.add(it["shipment"], it["values"]["주문상품명"], mid, dry_run)
            result["sent"] += len(batch)
        log.info("[%s] %s %d건 완료", name, "DRY-RUN" if dry_run else "발송", result["sent"])

    rp = out_dir / f"{shop_key}_결과_{stamp}.xlsx"
    write_report(rp, items, skipped)
    log.info("결과 리포트: %s", rp)
    result["report"] = str(rp)
    return result
